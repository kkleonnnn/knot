"""tests/api/test_exports.py — v0.4.0 CSV 导出端点集成测试。

覆盖：所有者导出成功 / 无消息 404 / 空 rows 400 / analyst 不可下载他人 message (404 防枚举) /
admin 可下载任何 message / 文件名 + Content-Type + utf-8-sig BOM。
"""
from __future__ import annotations


def _seed_message_for(client, headers, rows: list[dict]) -> tuple[int, int]:
    """创建 conv，写一条 message（直接调 repo），返回 (conv_id, message_id)。"""
    create = client.post("/api/conversations", json={"title": "导出测试"}, headers=headers)
    assert create.status_code == 200, create.text
    cid = create.json()["id"]

    from knot.repositories.message_repo import save_message
    mid = save_message(
        conv_id=cid, question="列出昨天注册用户", sql="SELECT id FROM users",
        explanation="示例", confidence="high",
        rows=rows, db_error="", cost_usd=0.0,
        input_tokens=0, output_tokens=0, retry_count=0,
        intent="detail",
    )
    return cid, mid


def test_export_csv_owner_success(client, auth_headers):
    """admin（也是 conv 所有者）可下载自己的 message。"""
    rows = [{"user_id": 1001, "name": "张三"}, {"user_id": 1002, "name": "李四"}]
    _conv_id, mid = _seed_message_for(client, auth_headers, rows)

    r = client.get(f"/api/messages/{mid}/export.csv", headers=auth_headers)
    assert r.status_code == 200, r.text
    # Content-Type 含 charset
    ct = r.headers.get("content-type", "")
    assert "text/csv" in ct
    # Content-Disposition 含文件名
    cd = r.headers.get("content-disposition", "")
    assert f"export_msg{mid}.csv" in cd
    # body 是 utf-8-sig (BOM 起手) + 中文 round-trip
    body = r.content
    assert body.startswith(b"\xef\xbb\xbf")
    text = body.decode("utf-8-sig")
    assert "张三" in text and "李四" in text and "user_id,name" in text


def test_export_csv_message_not_found_404(client, auth_headers):
    r = client.get("/api/messages/999999/export.csv", headers=auth_headers)
    assert r.status_code == 404


def test_export_csv_empty_rows_400(client, auth_headers):
    """rows 为空的 message → 400（无可导出数据），不是空 CSV 200。"""
    _cid, mid = _seed_message_for(client, auth_headers, rows=[])
    r = client.get(f"/api/messages/{mid}/export.csv", headers=auth_headers)
    assert r.status_code == 400


def test_export_csv_analyst_cannot_download_others_message(client, auth_headers):
    """非所有者 analyst 取他人 message_id → 404（防枚举），不是 403。"""
    # admin 创建一条带 rows 的 message
    rows = [{"x": 1}]
    _cid, admin_mid = _seed_message_for(client, auth_headers, rows)

    # 创建 analyst
    create = client.post(
        "/api/admin/users",
        json={"username": "bob", "password": "p", "role": "analyst"},
        headers=auth_headers,
    )
    assert create.status_code == 200
    login = client.post("/api/auth/login", json={"username": "bob", "password": "p"})
    analyst_headers = {"Authorization": f"Bearer {login.json()['token']}"}

    # analyst 直接拿 admin_mid 下载 → 404 防 message_id 枚举
    r = client.get(f"/api/messages/{admin_mid}/export.csv", headers=analyst_headers)
    assert r.status_code == 404


def test_export_csv_admin_can_download_any_message(client, auth_headers):
    """admin 角色 example：自己 conv 的 message 必然能下；admin 越权下他人 message 也允许。"""
    # 创 analyst + 用 analyst 创一条 message，admin 应能下
    create = client.post(
        "/api/admin/users",
        json={"username": "carol", "password": "p", "role": "analyst"},
        headers=auth_headers,
    )
    assert create.status_code == 200
    login = client.post("/api/auth/login", json={"username": "carol", "password": "p"})
    analyst_headers = {"Authorization": f"Bearer {login.json()['token']}"}

    rows = [{"x": 7}]
    _cid, analyst_mid = _seed_message_for(client, analyst_headers, rows)

    # admin 越权下载 → 200
    r = client.get(f"/api/messages/{analyst_mid}/export.csv", headers=auth_headers)
    assert r.status_code == 200
    assert r.content.startswith(b"\xef\xbb\xbf")


# ── v0.4.2: xlsx 导出 ────────────────────────────────────────────────────────


def test_export_xlsx_owner_success_with_metadata_headers(client, auth_headers):
    """xlsx 导出成功 + R-S7 响应头含 X-Export-Truncated/Total-Rows/Returned-Rows。"""
    rows = [{"用户": "张三", "金额": 100}, {"用户": "李四", "金额": 200}]
    _cid, mid = _seed_message_for(client, auth_headers, rows=rows)
    r = client.get(f"/api/messages/{mid}/export.xlsx", headers=auth_headers)
    assert r.status_code == 200
    # Content-Type
    ct = r.headers.get("content-type", "")
    assert "spreadsheet" in ct
    # R-S7 metadata headers
    assert r.headers.get("x-export-truncated") == "false"
    assert r.headers.get("x-export-total-rows") == "2"
    assert r.headers.get("x-export-returned-rows") == "2"
    # 文件名
    cd = r.headers.get("content-disposition", "")
    assert f"export_msg{mid}.xlsx" in cd
    # body 是合法 xlsx（用 openpyxl 反读）
    from io import BytesIO

    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(r.content)).active
    assert ws.cell(2, 1).value == "张三"
    assert ws.cell(2, 2).value == 100  # 数字保留 number 类型


def test_export_xlsx_truncation_metadata_when_over_5000_rows(client, auth_headers):
    """R-S7：>5000 行时 X-Export-Truncated=true，前端可据此 toast 提示。"""
    rows = [{"i": i} for i in range(5050)]
    _cid, mid = _seed_message_for(client, auth_headers, rows=rows)
    r = client.get(f"/api/messages/{mid}/export.xlsx", headers=auth_headers)
    assert r.status_code == 200
    assert r.headers.get("x-export-truncated") == "true"
    assert r.headers.get("x-export-total-rows") == "5050"
    assert r.headers.get("x-export-returned-rows") == "5000"


def test_export_xlsx_other_users_message_returns_404(client, auth_headers):
    """xlsx 同样守 message_id 枚举（与 csv 同款权限）。"""
    _cid, admin_mid = _seed_message_for(client, auth_headers, rows=[{"x": 1}])
    create = client.post(
        "/api/admin/users",
        json={"username": "dave", "password": "p", "role": "analyst"},
        headers=auth_headers,
    )
    assert create.status_code == 200
    login = client.post("/api/auth/login", json={"username": "dave", "password": "p"})
    analyst_headers = {"Authorization": f"Bearer {login.json()['token']}"}
    r = client.get(f"/api/messages/{admin_mid}/export.xlsx", headers=analyst_headers)
    assert r.status_code == 404


def test_export_xlsx_empty_rows_400(client, auth_headers):
    _cid, mid = _seed_message_for(client, auth_headers, rows=[])
    r = client.get(f"/api/messages/{mid}/export.xlsx", headers=auth_headers)
    assert r.status_code == 400
