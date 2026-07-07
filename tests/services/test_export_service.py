"""tests/services/test_export_service.py — v0.4.0 CSV 导出 service 单测。

覆盖：utf-8-sig BOM / 字段顺序 / 复杂值 JSON 序列化 / 空 rows / 中文字符。
"""
from __future__ import annotations

import csv
from io import StringIO

from knot.services.export_service import rows_to_csv_bytes


def _decode(b: bytes) -> str:
    """utf-8-sig 解码会自动剥掉 BOM；用 utf-8 解码可以验证 BOM 字节存在。"""
    return b.decode("utf-8-sig")


def test_empty_rows_returns_empty_bytes():
    assert rows_to_csv_bytes([]) == b""


def test_bom_byte_order_mark_present():
    """utf-8-sig 必须以 BOM (\xef\xbb\xbf) 起手，否则 Excel 中文乱码。"""
    out = rows_to_csv_bytes([{"a": 1}])
    assert out.startswith(b"\xef\xbb\xbf")


def test_basic_csv_roundtrip():
    rows = [{"name": "alice", "age": 30}, {"name": "bob", "age": 25}]
    out = rows_to_csv_bytes(rows)
    text = _decode(out)
    parsed = list(csv.DictReader(StringIO(text)))
    assert parsed == [{"name": "alice", "age": "30"}, {"name": "bob", "age": "25"}]


def test_chinese_chars_round_trip():
    rows = [{"用户": "张三", "金额": "¥100"}]
    out = rows_to_csv_bytes(rows)
    text = _decode(out)
    assert "用户" in text and "张三" in text and "¥100" in text


def test_explicit_cols_controls_order_and_subset():
    rows = [{"a": 1, "b": 2, "c": 3}]
    out = rows_to_csv_bytes(rows, cols=["c", "a"])
    text = _decode(out)
    header = text.splitlines()[0]
    assert header == "c,a"  # 顺序锁住，b 被丢弃


def test_complex_value_json_serialized():
    rows = [{"id": 1, "extra": {"nested": "中文"}}]
    out = rows_to_csv_bytes(rows)
    text = _decode(out)
    # dict 应被 JSON 序列化；csv writer 会把内层 " 双引号化（""..."")
    # 关键点：(1) 中文不被 ensure_ascii 转义；(2) JSON key/value 都进得去
    assert "中文" in text
    assert "nested" in text
    # 通过 csv.DictReader 反解，验证整个值是合法 JSON
    parsed = list(csv.DictReader(StringIO(text)))
    import json as _json
    assert _json.loads(parsed[0]["extra"]) == {"nested": "中文"}


def test_list_value_json_serialized():
    rows = [{"id": 1, "tags": ["a", "b", "c"]}]
    out = rows_to_csv_bytes(rows)
    text = _decode(out)
    assert "[" in text and "a" in text and "b" in text and "c" in text


def test_none_value_renders_empty():
    rows = [{"a": 1, "b": None}]
    out = rows_to_csv_bytes(rows)
    text = _decode(out)
    # CSV: "1," 表示 a=1, b=空
    data_line = text.splitlines()[1]
    assert data_line.endswith(",")  # b 字段值为空字符串


# ── v0.4.2 xlsx 导出 ────────────────────────────────────────────────────────


def test_xlsx_basic_round_trip():
    """xlsx 写入 → 用 openpyxl 读回，字段值一致；数字保留为 number 类型。"""
    from io import BytesIO

    from openpyxl import load_workbook

    from knot.services.export_service import rows_to_xlsx_bytes
    rows = [
        {"name": "alice", "age": 30, "score": 95.5},
        {"name": "bob", "age": 25, "score": 88.0},
    ]
    xlsx_bytes, meta = rows_to_xlsx_bytes(rows)
    assert meta["truncated"] is False
    assert meta["total"] == 2
    assert meta["exported"] == 2

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    # row 1 = header
    assert [c.value for c in ws[1]] == ["name", "age", "score"]
    # row 2/3 = data，数字保留为 number
    assert ws.cell(2, 1).value == "alice"
    assert ws.cell(2, 2).value == 30
    assert isinstance(ws.cell(2, 2).value, int)
    assert ws.cell(2, 3).value == 95.5
    assert isinstance(ws.cell(2, 3).value, float)


def test_xlsx_chinese_chars_preserved():
    """中文字段名 + 内容在 xlsx 中保留 unicode。"""
    from io import BytesIO

    from openpyxl import load_workbook

    from knot.services.export_service import rows_to_xlsx_bytes
    rows = [{"用户": "张三", "金额": 100}]
    xlsx_bytes, _ = rows_to_xlsx_bytes(rows)
    ws = load_workbook(BytesIO(xlsx_bytes)).active
    assert ws.cell(1, 1).value == "用户"
    assert ws.cell(2, 1).value == "张三"


def test_xlsx_truncates_at_5000_rows_with_metadata():
    """R-15 + R-S7：超过 5000 行截断，metadata 暴露 truncated=true / total / exported。"""
    from knot.services.export_service import XLSX_MAX_ROWS, rows_to_xlsx_bytes
    big_rows = [{"i": i} for i in range(XLSX_MAX_ROWS + 1234)]
    _, meta = rows_to_xlsx_bytes(big_rows)
    assert meta["truncated"] is True
    assert meta["total"] == XLSX_MAX_ROWS + 1234
    assert meta["exported"] == XLSX_MAX_ROWS


def test_xlsx_complex_value_json_serialized():
    """dict/list 复杂值 → JSON 字符串落 cell（不丢失数据）。"""
    from io import BytesIO
    import json as _json

    from openpyxl import load_workbook

    from knot.services.export_service import rows_to_xlsx_bytes
    rows = [{"id": 1, "extra": {"nested": "中文", "n": 42}}]
    xlsx_bytes, _ = rows_to_xlsx_bytes(rows)
    ws = load_workbook(BytesIO(xlsx_bytes)).active
    cell_v = ws.cell(2, 2).value
    assert isinstance(cell_v, str)
    parsed = _json.loads(cell_v)
    assert parsed == {"nested": "中文", "n": 42}


def test_xlsx_empty_rows_returns_metadata_zero():
    """空 rows → metadata 0 / 0 / not truncated；xlsx 文件仍合法（可打开但无数据）。"""
    from knot.services.export_service import rows_to_xlsx_bytes
    xlsx_bytes, meta = rows_to_xlsx_bytes([])
    assert meta == {"truncated": False, "total": 0, "exported": 0}
    assert xlsx_bytes  # bytes 非空（openpyxl 写了一个 sheet 头）


# ── v0.8.4 CSV / 公式注入中性化（安全 chore）────────────────────────────────

_DANGEROUS = [
    '=SUMIF(B4:B28, "USDT", O4:O28)',   # kk 截图的公式
    "+HYPERLINK(\"http://evil\")",
    "@SUM(1)",
    "-2+3",                              # formula-ish（非纯数字）
    "=cmd|'/c calc'!A1",                # 经典 CSV 注入 payload
    "\t=1+1",                           # 前导 TAB
    "\r=1+1",                           # 前导 CR
]


def test_csv_neutralizes_formula_injection_cells():
    """首字符 = + - @ TAB CR 的文本单元格导出前应前缀 ' 中性化。"""
    rows = [{"v": s} for s in _DANGEROUS]
    parsed = list(csv.DictReader(StringIO(_decode(rows_to_csv_bytes(rows)))))
    for i, s in enumerate(_DANGEROUS):
        assert parsed[i]["v"] == "'" + s, f"危险文本应前缀 ': {s!r} → {parsed[i]['v']!r}"


def test_csv_does_not_mangle_numbers():
    """数值（int/float）+ 纯数字字符串（含负数 / 科学计数）不应被中性化。"""
    rows = [{"i": -5, "f": -3.2, "s_neg": "-5", "s_sci": "-1.5e3", "s_pos": "+42"}]
    p = list(csv.DictReader(StringIO(_decode(rows_to_csv_bytes(rows)))))[0]
    assert p["i"] == "-5"
    assert p["f"] == "-3.2"
    assert p["s_neg"] == "-5"       # 纯数字字符串不加 '
    assert p["s_sci"] == "-1.5e3"
    assert p["s_pos"] == "+42"


def test_csv_neutralizes_dangerous_header():
    """列名（= SQL 别名，可控）也是注入向量 → 表头中性化；安全列名不变。"""
    rows = [{"=danger": 1, "safe": 2}]
    hdr = list(csv.reader(StringIO(_decode(rows_to_csv_bytes(rows)))))[0]
    assert hdr[0] == "'=danger"
    assert hdr[1] == "safe"


def test_csv_sumif_repro_from_screenshot():
    """kk 截图复现：文本单元格 =SUMIF(...) 导出不再作为公式执行（前缀 '）。"""
    s = '=SUMIF(B4:B28, "USDT", O4:O28)'
    p = list(csv.DictReader(StringIO(_decode(rows_to_csv_bytes([{"计算": s}])))))[0]
    assert p["计算"] == "'" + s


def test_xlsx_neutralizes_injection_and_keeps_numbers():
    """xlsx：文本单元格中性化；数值 native 不动；纯数字字符串不动；formula-ish 中性化。"""
    from io import BytesIO

    from openpyxl import load_workbook

    from knot.services.export_service import rows_to_xlsx_bytes
    rows = [{"txt": '=SUMIF(A1:A2,"x",B1:B2)', "neg": -5, "s_neg": "-5", "fx": "-2+3"}]
    ws = load_workbook(BytesIO(rows_to_xlsx_bytes(rows)[0])).active
    assert ws.cell(2, 1).value == '\'=SUMIF(A1:A2,"x",B1:B2)'  # 文本中性化
    assert ws.cell(2, 2).value == -5 and isinstance(ws.cell(2, 2).value, int)  # native number
    assert ws.cell(2, 3).value == "-5"      # 纯数字字符串不中性化
    assert ws.cell(2, 4).value == "'-2+3"   # formula-ish 中性化


def test_xlsx_neutralizes_dangerous_header():
    """xlsx 表头中性化。"""
    from io import BytesIO

    from openpyxl import load_workbook

    from knot.services.export_service import rows_to_xlsx_bytes
    ws = load_workbook(BytesIO(rows_to_xlsx_bytes([{"@evil": 1, "ok": 2}])[0])).active
    assert ws.cell(1, 1).value == "'@evil"
    assert ws.cell(1, 2).value == "ok"
