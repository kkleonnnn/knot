"""knot/api/admin/metrics.py — 指标注册表管理路由（v0.7.0 C2 语义层第一刀）。

route 前缀 `/api/admin/metrics-registry`（**避与既有 `/api/admin/metrics` 内测健康 KPI 屏撞**）。
全 require_tenant_admin（v0.7.0 metric 仅 admin 可见，继承 R-2FA enroll gate）。CRUD 经 metric_repo（OOS-1v2 死锁）。
审计接线：metric.create/update/delete（AuditAction）+ resource_type=metric。
"""
from __future__ import annotations

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from pydantic import BaseModel

from knot.api._audit_helpers import audit
from knot.api.deps import require_tenant_admin
from knot.models.errors import MetadataError
from knot.repositories import metric_repo

router = APIRouter()


class MetricCreateRequest(BaseModel):
    catalog_id: int = 1
    name: str
    caliber: str = ""              # 原子指标必填；派生指标（lineage）免 caliber（repo _validate 兜底）
    display: str = ""
    aliases: str = ""              # JSON list
    base_object: str = ""
    filters: str = ""              # JSON list
    dimensions: str = ""           # JSON list
    date_column: str = ""          # 时间窗注入列名（v0.7.17；显式优先，空=按维度名 regex 推断）
    lineage: str = ""              # 结构化派生定义 JSON {op,left,right}（v0.7.16 激活；空=原子）
    freshness_lag_days: int = 1
    enabled: int = 1


class MetricUpdateRequest(BaseModel):
    name: str | None = None
    caliber: str | None = None
    display: str | None = None
    aliases: str | None = None
    base_object: str | None = None
    filters: str | None = None
    dimensions: str | None = None
    date_column: str | None = None
    lineage: str | None = None
    freshness_lag_days: int | None = None
    enabled: int | None = None


@router.get("/api/admin/metrics-registry")
async def admin_list_metrics(catalog_id: int | None = None, admin=Depends(require_tenant_admin)):
    return metric_repo.list_metrics(catalog_id)


@router.get("/api/admin/metrics-registry/{metric_id}")
async def admin_get_metric(metric_id: int, admin=Depends(require_tenant_admin)):
    m = metric_repo.get_metric(metric_id)
    if m is None:
        raise HTTPException(status_code=404, detail="指标不存在")
    return m


@router.post("/api/admin/metrics-registry")
async def admin_create_metric(req: MetricCreateRequest, request: Request, admin=Depends(require_tenant_admin)):
    payload = req.dict()
    catalog_id = payload.pop("catalog_id", 1)
    try:
        mid = metric_repo.create_metric(catalog_id=catalog_id, **payload)
    except MetadataError as e:
        raise HTTPException(status_code=400, detail=str(e))
    audit(request, admin, action="metric.create", resource_type="metric", resource_id=mid,
          detail={"name": req.name, "catalog_id": catalog_id})
    return {"id": mid}


@router.put("/api/admin/metrics-registry/{metric_id}")
async def admin_update_metric(metric_id: int, req: MetricUpdateRequest, request: Request, admin=Depends(require_tenant_admin)):
    fields = {k: v for k, v in req.dict().items() if v is not None}
    try:
        metric_repo.update_metric(metric_id, **fields)
    except MetadataError as e:
        raise HTTPException(status_code=400, detail=str(e))
    audit(request, admin, action="metric.update", resource_type="metric", resource_id=metric_id,
          detail={"fields": sorted(fields.keys())})
    return {"ok": True}


@router.delete("/api/admin/metrics-registry/{metric_id}")
async def admin_delete_metric(metric_id: int, request: Request, admin=Depends(require_tenant_admin)):
    metric_repo.delete_metric(metric_id)
    audit(request, admin, action="metric.delete", resource_type="metric", resource_id=metric_id)
    return {"ok": True}


@router.post("/api/admin/metrics-registry/upload")
async def admin_upload_metrics(file: UploadFile = File(...), request: Request = None,
                               catalog_id: int = 1, admin=Depends(require_tenant_admin)):
    """v0.8.13 xlsx 批量导入指标（列 name/display/caliber/base_object/date_column/unit/aliases；
    aliases 逗号分隔 → JSON list）。逐行 create（当前 catalog；默认 #1）；行错不中断、收集返回。"""
    import json
    from io import BytesIO

    # v0.8.13 fixup：只收 .xlsx（openpyxl 读不了 legacy BIFF .xls；此前收 .xls 会过闸后死在 load_workbook 给误导性「解析失败」）
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 xlsx 文件")
    try:
        from openpyxl import load_workbook
        rows = list(load_workbook(filename=BytesIO(await file.read()), data_only=True).active.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)[:200]}")
    if not rows:
        raise HTTPException(status_code=400, detail="文件内容为空")
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    # v0.8.13 fixup：校验必需表头列 name（否则错文件/无表头会静默返回 {inserted:0, errors:[]} 无诊断）
    if "name" not in header:
        raise HTTPException(status_code=400, detail="表头缺少必需列「name」（请使用下载的模板表头）")

    def cell(d, k):
        v = d.get(k)
        return str(v).strip() if v is not None else ""

    inserted, errors = 0, []
    for r in rows[1:]:
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        name = cell(d, "name")
        if not name:
            continue
        al = [a.strip() for a in cell(d, "aliases").split(",") if a.strip()]
        try:
            metric_repo.create_metric(
                catalog_id=catalog_id, name=name, display=cell(d, "display"), caliber=cell(d, "caliber"),
                base_object=cell(d, "base_object"), date_column=cell(d, "date_column"), unit=cell(d, "unit"),
                aliases=json.dumps(al, ensure_ascii=False) if al else "",
            )
            inserted += 1
        except Exception as e:
            errors.append(f"{name}: {str(e)[:80]}")
    if inserted:
        audit(request, admin, action="metric.create", resource_type="metric",
              detail={"bulk_import": inserted, "catalog_id": catalog_id})
    return {"inserted": inserted, "errors": errors[:20]}
