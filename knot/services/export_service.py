"""export_service — v0.4.0 CSV / v0.4.2 xlsx 导出。

设计选择（手册 §4.1 + §3.5）：内存 BytesIO 模式。
- MAX_RESULT_ROWS=500 锁死单次 CSV 结果集 → ≤ 200KB
- xlsx：5000 行硬限（资深 R-15）；超出截断 + 守护者 R-S7 metadata 暴露给前端
- 中文 utf-8-sig（带 BOM）保证 Excel 直接打开不乱码（CSV 模式）
- xlsx 模式下 openpyxl 自动 utf-8 + 数字格式保留
"""
from __future__ import annotations

import csv
import json
import re
from io import BytesIO, StringIO

# v0.4.2 R-15 + R-S7：xlsx 单文件硬限 5000 行
XLSX_MAX_ROWS: int = 5000

# v0.8.4 安全 chore（CSV/公式注入中性化）：电子表格（Excel/Sheets/WPS）打开导出文件时，
# 首字符为 = + - @ TAB CR 的**文本**单元格会被当公式执行（formula / CSV injection）——
# 任何用户 / admin 可控文本进导出即为向量。对这类文本前缀 '（强制按文本处理）中性化。
# 纯数字字面（-5 / +3.2 / -1.5e3）不动，避免破坏合法负数 / 科学计数显示。
_INJECTION_PREFIXES: tuple[str, ...] = ("=", "+", "-", "@", "\t", "\r")
_NUMERIC_RE = re.compile(r"^[+-]?(\d+\.?\d*|\.\d+)([eE][+-]?\d+)?$")


def _neutralize_text(s: str) -> str:
    """对可能触发电子表格公式注入的**文本**单元格前缀 ' 中性化（v0.8.4）。

    - 首字符 ∈ {= + - @ TAB CR} → 前缀 '（Excel 视为纯文本，不求值）
    - 例外：首字符为 + / - 且整串是纯数字字面（-5 / +3.2 / -1.5e3）→ 不动（合法数值）
    - 其他文本原样返回
    """
    if s and s[0] in _INJECTION_PREFIXES:
        if s[0] in "+-" and _NUMERIC_RE.match(s):
            return s
        return "'" + s
    return s


def rows_to_csv_bytes(rows: list[dict], cols: list[str] | None = None,
                     headers: list[str] | None = None) -> bytes:
    """把 [{col: val}, ...] 转成 CSV 字节流（utf-8-sig，带 BOM）。

    cols=None 时从首行 keys 推断字段顺序。
    headers（v0.8.9）：表头显示名（默认 = cols）—— 值仍按 cols(key) 取，但表头行写 headers（如 BI 列中文 label）。
    复杂值（dict / list）使用 JSON 序列化（中文不转义）。
    空 rows 返回空 bytes。
    """
    if not rows:
        return b""
    cols = cols or list(rows[0].keys())
    hdr = headers if (headers and len(headers) == len(cols)) else cols
    sio = StringIO()
    writer = csv.DictWriter(sio, fieldnames=cols, extrasaction="ignore")
    # v0.8.4：表头也是注入向量（列名 = SQL 别名可控）→ 中性化后写（等价 writeheader() 对安全列名）
    writer.writerow({c: _neutralize_text(str(h)) for c, h in zip(cols, hdr)})
    for r in rows:
        writer.writerow({c: _stringify(r.get(c)) for c in cols})
    return sio.getvalue().encode("utf-8-sig")


def _safe_sheet_name(name: str, used: set) -> str:
    """Excel sheet 名约束（v0.8.9 多 sheet 导出）：≤31 字符 · 去 []:*?/\\ · 非空 · 去重。"""
    s = "".join(c for c in str(name or "") if c not in '[]:*?/\\')[:31].strip() or "Sheet"
    base, n = s, 1
    while s in used:
        suffix = f"~{n}"
        s = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(s)
    return s


def sheets_to_xlsx_bytes(sheets: list[dict]) -> tuple[bytes, dict]:
    """v0.8.9：多 sheet xlsx（tabbed 报表每页一 sheet）。sheets=[{name, rows, cols?, headers?}]。
    返回 (bytes, meta)；meta={"sheets":[{name,truncated,total,exported}], "truncated":bool}。中性化 + 5000 行/表硬限沿用。"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)                       # 去默认空 sheet
    used, metas = set(), []
    for i, sh in enumerate(sheets):
        rows = sh.get("rows") or []
        truncated = len(rows) > XLSX_MAX_ROWS
        actual = rows[:XLSX_MAX_ROWS] if truncated else rows
        cols = sh.get("cols") or (list(actual[0].keys()) if actual else [])
        hdr = sh.get("headers") if (sh.get("headers") and len(sh["headers"]) == len(cols)) else cols
        ws = wb.create_sheet(title=_safe_sheet_name(sh.get("name") or f"Sheet{i + 1}", used))
        if cols:
            ws.append([_neutralize_text(str(h)) for h in hdr])
        for r in actual:
            ws.append([_xlsx_value(r.get(c)) for c in cols])
        metas.append({"name": ws.title, "truncated": truncated, "total": len(rows), "exported": len(actual)})
    if not wb.sheetnames:                      # 全空 → 至少留一个 sheet 防 openpyxl save 崩
        wb.create_sheet(title="Empty")
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), {"sheets": metas, "truncated": any(m["truncated"] for m in metas)}


def rows_to_xlsx_bytes(
    rows: list[dict],
    cols: list[str] | None = None,
    sheet_name: str = "Data",
    headers: list[str] | None = None,
) -> tuple[bytes, dict]:
    """v0.4.2：rows → xlsx bytes + R-S7 metadata。

    返回 (xlsx_bytes, metadata)：
      metadata = {
          "truncated": bool,        # rows 是否被截断
          "total": int,             # 原始行数
          "exported": int,          # 实际写入 xlsx 的行数（≤ XLSX_MAX_ROWS）
      }

    资深 R-15：5000 行硬限防 OOM。
    守护者 R-S7：metadata 由 API 层放 response header（X-Export-*），前端 toast 提示。

    数字保留为 number 类型（Excel 自动右对齐 + 公式可用）；
    复杂值（dict/list）JSON 序列化为字符串。
    """
    from openpyxl import Workbook  # 延迟 import，避免不用 xlsx 时也加载
    truncated = len(rows) > XLSX_MAX_ROWS
    actual_rows = rows[:XLSX_MAX_ROWS] if truncated else rows
    cols = cols or (list(actual_rows[0].keys()) if actual_rows else [])

    hdr = headers if (headers and len(headers) == len(cols)) else cols
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if cols:
        ws.append([_neutralize_text(str(h)) for h in hdr])  # v0.8.4：表头中性化（v0.8.9 headers=显示名）
    for r in actual_rows:
        ws.append([_xlsx_value(r.get(c)) for c in cols])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), {
        "truncated": truncated,
        "total": len(rows),
        "exported": len(actual_rows),
    }


def _stringify(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)
    if isinstance(v, str):
        return _neutralize_text(v)  # v0.8.4：文本单元格注入中性化
    return str(v)  # int/float/bool 等数值 — 安全，不中性化（负数原样）


def _xlsx_value(v):
    """xlsx 写入值的类型转换。
    - None → '' （空 cell）
    - int/float → 保留为 number（Excel 自动识别）
    - bool → True/False（openpyxl 原生支持）
    - dict/list → JSON 字符串
    - 其他 → str()"""
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)
    return _neutralize_text(str(v))  # v0.8.4：文本单元格注入中性化
