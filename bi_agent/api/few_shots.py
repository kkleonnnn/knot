"""
few_shots.py — admin 维护 few-shot 示例（DB 存储）
"""
from fastapi import APIRouter, Body, Depends, File, HTTPException, Request, UploadFile

# v0.3.0: import persistence → 直接 import 各 repo（保留"persistence.X"调用形态）
from bi_agent.api._audit_helpers import audit
from bi_agent.api.deps import require_admin
from bi_agent.repositories import few_shot_repo

router = APIRouter()


@router.get("/api/few-shots")
async def list_few_shots(admin=Depends(require_admin)):
    return few_shot_repo.list_few_shots()


@router.post("/api/few-shots")
async def create_few_shot(payload: dict = Body(...), request: Request = None, admin=Depends(require_admin)):
    q = (payload.get("question") or "").strip()
    s = (payload.get("sql") or "").strip()
    t = (payload.get("type") or "").strip()
    if not q or not s:
        raise HTTPException(status_code=400, detail="question / sql 不能为空")
    fid = few_shot_repo.create_few_shot(q, s, t, 1 if payload.get("is_active", 1) else 0)
    audit(request, admin, action="config.few_shots_change", resource_type="few_shots",
          resource_id=fid, detail={"op": "create", "type": t})
    return {"id": fid}


@router.put("/api/few-shots/{fid}")
async def update_few_shot(fid: int, payload: dict = Body(...), request: Request = None, admin=Depends(require_admin)):
    fields = {k: v for k, v in payload.items() if k in {"question", "sql", "type", "is_active"}}
    few_shot_repo.update_few_shot(fid, **fields)
    audit(request, admin, action="config.few_shots_change", resource_type="few_shots",
          resource_id=fid, detail={"op": "update", "fields": sorted(fields.keys())})
    return {"ok": True}


@router.delete("/api/few-shots/{fid}")
async def delete_few_shot(fid: int, request: Request, admin=Depends(require_admin)):
    few_shot_repo.delete_few_shot(fid)
    audit(request, admin, action="config.few_shots_change", resource_type="few_shots",
          resource_id=fid, detail={"op": "delete"})
    return {"ok": True}


@router.post("/api/few-shots/upload")
async def upload_few_shots(file: UploadFile = File(...), admin=Depends(require_admin)):
    """xlsx：列名 question / sql / type（可选）/ is_active（可选）"""
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 xlsx 文件")
    try:
        from io import BytesIO

        from openpyxl import load_workbook
        data = await file.read()
        wb = load_workbook(filename=BytesIO(data), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="文件内容为空")
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        items = []
        for r in rows[1:]:
            d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            q = (d.get("question") or "").strip() if isinstance(d.get("question"), str) else (str(d.get("question")) if d.get("question") is not None else "")
            s = (d.get("sql") or "").strip() if isinstance(d.get("sql"), str) else (str(d.get("sql")) if d.get("sql") is not None else "")
            if not q or not s:
                continue
            t = d.get("type") or ""
            t = str(t).strip()
            ia = d.get("is_active")
            ia = 1 if ia is None or str(ia).strip() in ("1", "true", "True", "是", "yes") else (0 if str(ia).strip() in ("0", "false", "False", "否", "no") else 1)
            items.append({"question": q, "sql": s, "type": t, "is_active": ia})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)[:200]}")

    n = few_shot_repo.bulk_insert_few_shots(items)
    return {"inserted": n}
