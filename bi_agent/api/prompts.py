"""
prompts.py — admin 维护 3 个 agent 的 system prompt 覆盖
agent_name ∈ {clarifier, sql_planner, presenter}
v0.2.2: validator 已移除；仍存在的 validator 模板会被忽略（保留 DB 数据待清理）
"""
from fastapi import APIRouter, Body, Depends, File, HTTPException, Request, UploadFile

# v0.3.0: import persistence → 直接 import 各 repo（保留"persistence.X"调用形态）
from bi_agent.api._audit_helpers import audit
from bi_agent.api.deps import require_admin
from bi_agent.repositories import prompt_repo

router = APIRouter()

VALID_AGENTS = {"clarifier", "sql_planner", "presenter"}


@router.get("/api/prompts")
async def list_prompts(admin=Depends(require_admin)):
    rows = prompt_repo.list_prompt_templates()
    return rows


@router.get("/api/prompts/{agent_name}")
async def get_prompt(agent_name: str, admin=Depends(require_admin)):
    if agent_name not in VALID_AGENTS:
        raise HTTPException(status_code=404, detail="未知 agent")
    return {"agent_name": agent_name, "content": prompt_repo.get_prompt_template(agent_name)}


@router.put("/api/prompts/{agent_name}")
async def set_prompt(agent_name: str, payload: dict = Body(...), request: Request = None, admin=Depends(require_admin)):
    if agent_name not in VALID_AGENTS:
        raise HTTPException(status_code=404, detail="未知 agent")
    content = payload.get("content", "")
    prompt_repo.set_prompt_template(agent_name, content, updated_by=admin["id"])
    audit(request, admin, action="config.prompt_update", resource_type="prompt",
          resource_id=agent_name, detail={"op": "update", "content_len": len(content)})
    return {"ok": True}


@router.delete("/api/prompts/{agent_name}")
async def delete_prompt(agent_name: str, request: Request, admin=Depends(require_admin)):
    if agent_name not in VALID_AGENTS:
        raise HTTPException(status_code=404, detail="未知 agent")
    prompt_repo.delete_prompt_template(agent_name)
    audit(request, admin, action="config.prompt_update", resource_type="prompt",
          resource_id=agent_name, detail={"op": "delete"})
    return {"ok": True}


@router.post("/api/prompts/upload")
async def upload_prompts(file: UploadFile = File(...), admin=Depends(require_admin)):
    """xlsx：列名 agent_name / content"""
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 xlsx 文件")
    try:
        from io import BytesIO

        from openpyxl import load_workbook
        data = await file.read()
        wb = load_workbook(filename=BytesIO(data), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="文件内容为空")
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        n = 0
        for r in rows[1:]:
            d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            agent = (d.get("agent_name") or "").strip() if isinstance(d.get("agent_name"), str) else ""
            content = d.get("content") or ""
            content = str(content) if content is not None else ""
            if agent in VALID_AGENTS and content.strip():
                prompt_repo.set_prompt_template(agent, content, updated_by=admin["id"])
                n += 1
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败: {str(e)[:200]}")

    return {"updated": n}
